from flask import Flask, render_template, request, redirect
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)
FILE_NAME = "123.xlsx"

def create_file():
    if not os.path.exists(FILE_NAME):
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Age", "Sex", "Address"])
        wb.save(FILE_NAME)

@app.route("/", methods=["GET", "POST"])
def form():
    if request.method == "POST":
        names = request.form.getlist("name")
        ages = request.form.getlist("age")
        sexes = request.form.getlist("sex")
        addresses = request.form.getlist("address")

        wb = load_workbook(FILE_NAME)
        ws = wb.active

        for i in range(len(names)):
            ws.append([names[i], ages[i], sexes[i], addresses[i]])

        wb.save(FILE_NAME)
        return redirect("/success")

    return render_template("index.html")

@app.route("/success")
def success():
    return "<h2>Data saved successfully!</h2><a href='/'>Back</a>"

if __name__ == "__main__":
    create_file()
    app.run(host="0.0.0.0", port=5000, debug=True)